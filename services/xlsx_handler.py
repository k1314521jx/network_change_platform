"""XLSX 导入/导出工具函数"""
import json
from io import BytesIO
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, Alignment


# 表格列定义（严格顺序）
TABLE1_COLUMNS = ['row_index', 'raw_cmd', 'raw_rollback', 'step_name', 'au_name', 'role', 'entity', 'parameters']
TABLE2_COLUMNS = ['id', 'label', 'name', 'properties']
TABLE3_COLUMNS = ['source_entity', 'relation_type', 'target_entity', 'relation_attributes']

# 对象类型字段
OBJECT_FIELDS = {'parameters', 'properties', 'relation_attributes'}


def export_triple_to_xlsx(triple_json: dict) -> BytesIO:
    """
    将三元组 JSON 数据导出为 xlsx 字节流

    Args:
        triple_json: 包含 Table1_Alignment, Table2_Entities_Attributes, Table3_Relations 的字典

    Returns:
        BytesIO: xlsx 文件字节流
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    # 创建 3 个 sheet
    tables = [
        ('Table1_Alignment', TABLE1_COLUMNS, triple_json.get('Table1_Alignment', [])),
        ('Table2_Entities_Attributes', TABLE2_COLUMNS, triple_json.get('Table2_Entities_Attributes', [])),
        ('Table3_Relations', TABLE3_COLUMNS, triple_json.get('Table3_Relations', [])),
    ]

    for sheet_name, columns, data in tables:
        ws = wb.create_sheet(title=sheet_name)

        # 写入表头
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写入数据行
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, '')

                # 对象字段序列化为 JSON 字符串
                if col_name in OBJECT_FIELDS and isinstance(value, dict):
                    value = json.dumps(value, ensure_ascii=False)
                elif value is None:
                    value = ''

                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动调整列宽
        for col_idx, col_name in enumerate(columns, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    # 保存到字节流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_xlsx_to_triple(file_stream) -> dict:
    """
    从 xlsx 文件流解析三元组数据

    Args:
        file_stream: 文件流对象（Flask request.files['file']）

    Returns:
        dict: 包含 Table1_Alignment, Table2_Entities_Attributes, Table3_Relations 的字典

    Raises:
        ValueError: 文件格式不符合要求时抛出异常
    """
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
    except Exception as e:
        raise ValueError(f"无法读取 xlsx 文件: {str(e)}")

    # 验证 sheet 数量和名称
    expected_sheets = ['Table1_Alignment', 'Table2_Entities_Attributes', 'Table3_Relations']
    if len(wb.sheetnames) != 3:
        raise ValueError(f"xlsx 文件必须包含 3 个 sheet，实际包含 {len(wb.sheetnames)} 个")

    for idx, expected_name in enumerate(expected_sheets):
        actual_name = wb.sheetnames[idx]
        if actual_name != expected_name:
            raise ValueError(
                f"第 {idx + 1} 个 sheet 名称错误：期望 '{expected_name}'，实际 '{actual_name}'"
            )

    result = {}

    # 解析 Table1
    result['Table1_Alignment'] = _parse_table1(wb['Table1_Alignment'])

    # 解析 Table2
    result['Table2_Entities_Attributes'] = _parse_table2(wb['Table2_Entities_Attributes'])

    # 解析 Table3
    result['Table3_Relations'] = _parse_table3(wb['Table3_Relations'])

    return result


def _parse_table1(ws) -> List[Dict[str, Any]]:
    """解析 Table1_Alignment sheet"""
    # 验证列顺序
    header_row = [cell.value for cell in ws[1]]
    if header_row != TABLE1_COLUMNS:
        raise ValueError(
            f"Table1_Alignment 列顺序错误。\n"
            f"期望: {TABLE1_COLUMNS}\n"
            f"实际: {header_row}"
        )

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, col_name in enumerate(TABLE1_COLUMNS, start=1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value

            # 处理空值
            if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ''):
                if col_name == 'parameters':
                    row_data[col_name] = {}
                else:
                    row_data[col_name] = ''
            elif col_name == 'parameters':
                # 反序列化 JSON
                try:
                    row_data[col_name] = json.loads(str(cell_value))
                    if not isinstance(row_data[col_name], dict):
                        raise ValueError(f"parameters 必须是对象")
                except json.JSONDecodeError:
                    raise ValueError(
                        f"Table1_Alignment 第 {row_idx} 行 parameters 字段不是有效的 JSON: {cell_value}"
                    )
            elif col_name == 'row_index':
                # row_index 转为整数
                try:
                    row_data[col_name] = int(cell_value) if cell_value else row_idx - 1
                except (ValueError, TypeError):
                    row_data[col_name] = row_idx - 1
            else:
                row_data[col_name] = str(cell_value).strip()

        # 验证必需字段（raw_rollback 可选）
        required_fields = {'row_index', 'raw_cmd', 'step_name', 'au_name', 'role', 'entity', 'parameters'}
        missing_fields = required_fields - set(row_data.keys())
        if missing_fields:
            raise ValueError(f"Table1_Alignment 第 {row_idx} 行缺少必需字段: {missing_fields}")

        # 验证没有多余字段
        extra_fields = set(row_data.keys()) - set(TABLE1_COLUMNS)
        if extra_fields:
            raise ValueError(f"Table1_Alignment 第 {row_idx} 行包含多余字段: {extra_fields}")

        rows.append(row_data)

    return rows


def _parse_table2(ws) -> List[Dict[str, Any]]:
    """解析 Table2_Entities_Attributes sheet"""
    # 验证列顺序
    header_row = [cell.value for cell in ws[1]]
    if header_row != TABLE2_COLUMNS:
        raise ValueError(
            f"Table2_Entities_Attributes 列顺序错误。\n"
            f"期望: {TABLE2_COLUMNS}\n"
            f"实际: {header_row}"
        )

    valid_labels = {"Scenario", "LogicalStep", "ActionUnit", "CLITemplate", "NetworkEntity", "Parameter"}

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, col_name in enumerate(TABLE2_COLUMNS, start=1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value

            # 处理空值
            if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ''):
                if col_name == 'properties':
                    row_data[col_name] = {}
                else:
                    row_data[col_name] = ''
            elif col_name == 'properties':
                # 反序列化 JSON
                try:
                    row_data[col_name] = json.loads(str(cell_value))
                    if not isinstance(row_data[col_name], dict):
                        raise ValueError(f"properties 必须是对象")
                except json.JSONDecodeError:
                    raise ValueError(
                        f"Table2_Entities_Attributes 第 {row_idx} 行 properties 字段不是有效的 JSON: {cell_value}"
                    )
            else:
                row_data[col_name] = str(cell_value).strip()

        # 验证 label 枚举值
        if row_data.get('label') not in valid_labels:
            raise ValueError(
                f"Table2_Entities_Attributes 第 {row_idx} 行 label 值非法: '{row_data.get('label')}'，"
                f"必须是 {valid_labels} 之一"
            )

        # 验证必需字段
        required_fields = {'id', 'label', 'name', 'properties'}
        missing_fields = required_fields - set(row_data.keys())
        if missing_fields:
            raise ValueError(f"Table2_Entities_Attributes 第 {row_idx} 行缺少必需字段: {missing_fields}")

        # 验证没有多余字段
        extra_fields = set(row_data.keys()) - set(TABLE2_COLUMNS)
        if extra_fields:
            raise ValueError(f"Table2_Entities_Attributes 第 {row_idx} 行包含多余字段: {extra_fields}")

        rows.append(row_data)

    return rows


def _parse_table3(ws) -> List[Dict[str, Any]]:
    """解析 Table3_Relations sheet"""
    # 验证列顺序
    header_row = [cell.value for cell in ws[1]]
    if header_row != TABLE3_COLUMNS:
        raise ValueError(
            f"Table3_Relations 列顺序错误。\n"
            f"期望: {TABLE3_COLUMNS}\n"
            f"实际: {header_row}"
        )

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, col_name in enumerate(TABLE3_COLUMNS, start=1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value

            # 处理空值
            if cell_value is None or (isinstance(cell_value, str) and cell_value.strip() == ''):
                if col_name == 'relation_attributes':
                    row_data[col_name] = {}
                else:
                    row_data[col_name] = ''
            elif col_name == 'relation_attributes':
                # 反序列化 JSON
                try:
                    row_data[col_name] = json.loads(str(cell_value))
                    if not isinstance(row_data[col_name], dict):
                        raise ValueError(f"relation_attributes 必须是对象")
                except json.JSONDecodeError:
                    raise ValueError(
                        f"Table3_Relations 第 {row_idx} 行 relation_attributes 字段不是有效的 JSON: {cell_value}"
                    )
            else:
                row_data[col_name] = str(cell_value).strip()

        # 验证必需字段
        required_fields = {'source_entity', 'relation_type', 'target_entity', 'relation_attributes'}
        missing_fields = required_fields - set(row_data.keys())
        if missing_fields:
            raise ValueError(f"Table3_Relations 第 {row_idx} 行缺少必需字段: {missing_fields}")

        # 验证没有多余字段
        extra_fields = set(row_data.keys()) - set(TABLE3_COLUMNS)
        if extra_fields:
            raise ValueError(f"Table3_Relations 第 {row_idx} 行包含多余字段: {extra_fields}")

        rows.append(row_data)

    return rows
