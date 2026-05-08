import re
import openpyxl

DEVICE_PATTERN = re.compile(r'^[A-Za-z0-9-]+-\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
SECTION_TYPES = {'Pre-check': 'pre_check', 'Command': 'exe_command', 'Post-check': 'post_check'}
SKIP_COMMANDS = {'y', 'n', 'Y', 'N'}


def is_device(text):
    if not text:
        return False
    return bool(DEVICE_PATTERN.match(text.strip()))


def parse_demand_info(ws):
    device_info = []
    implementation_step = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if len(row) < 2 or row[1] is None:
            continue
        text = str(row[1]).strip()
        if not text:
            continue
        if is_device(text):
            device_info.append(text)
        else:
            implementation_step.append(text)
    return {"device_info": device_info, "implementation_step": implementation_step}


def parse_config(ws):
    config = {}
    current_device = None
    current_commands = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if len(row) < 1 or row[0] is None:
            continue
        text = str(row[0]).strip()
        if not text:
            continue
        if is_device(text):
            if current_device:
                config[current_device] = current_commands
            current_device = text
            current_commands = []
        elif text.startswith('<') and '>' in text:
            continue
        elif text == '#':
            continue
        elif text.startswith('Error:'):
            continue
        else:
            current_commands.append(text)
    if current_device:
        config[current_device] = current_commands
    return config


def parse_step_info(ws):
    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        c = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
        d = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
        e = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
        rows_data.append((b, c, d, e))

    if rows_data and 'step type' in rows_data[0][0].lower():
        rows_data = rows_data[1:]

    big_steps = []
    current_big_step = None
    current_section = None
    current_item = None

    def finalize_item():
        nonlocal current_item
        if current_item and current_section:
            if current_item['command']:
                current_section['items'].append(current_item)
        current_item = None

    def finalize_section():
        nonlocal current_section
        finalize_item()
        if current_section and current_big_step:
            current_big_step['sections'].append(current_section)
        current_section = None

    def finalize_big_step():
        nonlocal current_big_step
        finalize_section()
        if current_big_step:
            big_steps.append(current_big_step)
        current_big_step = None

    for b, c, d, e in rows_data:
        if b and not c and not e and b not in SECTION_TYPES:
            finalize_big_step()
            current_big_step = {'name': b, 'sections': []}
            continue

        if b in SECTION_TYPES:
            finalize_section()
            devices = []
            if c and is_device(c):
                devices.append(c)
            current_section = {
                'type': SECTION_TYPES[b],
                'devices': devices,
                'items': []
            }
            current_item = None
            continue

        if not current_section:
            continue

        if not b and not c and not d and not e:
            finalize_item()
            continue

        if c and is_device(c):
            current_section['devices'].append(c)
            continue

        if current_section['type'] in ('pre_check', 'post_check'):
            if c:
                finalize_item()
                current_item = {
                    'command': [c],
                    'rollback_command': [],
                    'desc': d if d else ''
                }
        elif current_section['type'] == 'exe_command':
            if c and c in SKIP_COMMANDS:
                continue
            if c and e:
                finalize_item()
                current_item = {
                    'command': [c],
                    'rollback_command': [e],
                    'desc': d if d else ''
                }
            elif c and not e:
                if current_item is None:
                    current_item = {
                        'command': [c],
                        'rollback_command': [],
                        'desc': d if d else ''
                    }
                else:
                    current_item['command'].append(c)
            elif e and not c:
                if current_item:
                    current_item['rollback_command'].append(e)

    finalize_big_step()

    step_info = []
    for big_step in big_steps:
        all_devices = []
        seen_devices = set()
        for section in big_step['sections']:
            for dev in section['devices']:
                if dev not in seen_devices:
                    all_devices.append(dev)
                    seen_devices.add(dev)

        sections_by_type = {'pre_check': [], 'exe_command': [], 'post_check': []}
        for section in big_step['sections']:
            sections_by_type[section['type']].append(section)

        def merge_items(sections):
            if not sections:
                return []
            return sections[0]['items']

        step = {
            'step_index': len(step_info) + 1,
            'step_name': big_step['name'],
            'command': {
                'pre_check': merge_items(sections_by_type['pre_check']),
                'exe_command': merge_items(sections_by_type['exe_command']),
                'post_check': merge_items(sections_by_type['post_check']),
            },
            'device': all_devices,
        }
        step_info.append(step)

    return step_info


def parse_xlsx(file_path):
    """解析Excel文件，返回结构化JSON"""
    import os
    wb = openpyxl.load_workbook(file_path)
    demand_info = parse_demand_info(wb.worksheets[0])
    config = parse_config(wb.worksheets[1])
    step_info = parse_step_info(wb.worksheets[2])
    name = os.path.splitext(os.path.basename(file_path))[0]
    return {
        'name': name,
        'demand_info': demand_info,
        'config': config,
        'step_info': step_info,
    }
